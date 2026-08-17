"""
라벨링용 엑셀 템플릿 생성기.

실행:
    python generate_label_template.py

결과:
    stage1_labeled.xlsx  (지식 안내 / 멤버십 혜택 안내 — 2클래스)
    stage2_labeled.xlsx  (멤버십 혜택 세부 — 8클래스)

두 파일 모두 이 스크립트가 있는 디렉터리에 생성됩니다.
이미 파일이 존재하면 덮어쓰지 않고 건너뜁니다 (기존 라벨링 데이터 보호).
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

THIS_DIR = os.path.dirname(os.path.abspath(__file__))
FONT_NAME = "Arial"

STAGE1_LABELS = ["knowledge", "membership"]
STAGE1_EXAMPLES = [
    ("요금제 종류가 어떻게 되나요?", "knowledge", ""),
    ("이번 달 생일 혜택 뭐 있어요?", "membership", ""),
]

STAGE2_LABELS = [
    "vip_choice",
    "multi_condition",
    "point_abolition",
    "grade_info",
    "birthday_benefit",
    "dal_benefit",
    "membership_type",
    "usage_history",
]
STAGE2_EXAMPLES = [
    ("VIP 초이스 이번 달에 몇 번 남았어?", "vip_choice", ""),
    ("VIP 등급이고 이번 달 생일인데 어디서 혜택 받을 수 있어?", "multi_condition", "생일+등급 조건 결합 → multi_condition"),
    ("포인트 없어진다는 게 무슨 말이에요?", "point_abolition", ""),
    ("제 등급이 뭐예요?", "grade_info", ""),
    ("생일 혜택만 알려줘요", "birthday_benefit", "조건 결합 없이 생일만 단독으로 묻는 경우"),
    ("이번 달 달달 혜택 뭐예요?", "dal_benefit", ""),
    ("멤버십 종류 뭐가 있어요?", "membership_type", ""),
    ("제가 예전에 뭐 썼는지 보여줘요", "usage_history", ""),
]

HEADERS = ["id", "utterance", "label", "ambiguous(O/X)", "labeler", "date_labeled"]
LEGEND = (
    "작성 방법: utterance에 실제(또는 예상) 사용자 발화를 입력하고, "
    "label 컬럼은 드롭다운에서 선택하세요. 판단이 애매했던 경우 ambiguous에 O 표시."
)


def _build_sheet(ws, labels: list[str], examples: list[tuple[str, str, str]]):
    ws.sheet_view.showGridLines = True

    # 안내 문구
    ws["A1"] = LEGEND
    ws["A1"].font = Font(name=FONT_NAME, italic=True, size=9, color="666666")
    ws.merge_cells("A1:F1")

    header_row = 3
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 예시 행 (연한 노란색 배경으로 "예시"임을 표시)
    example_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    row = header_row + 1
    for i, (utterance, label, note) in enumerate(examples, start=1):
        values = [f"EX-{i}", utterance, label, note, "예시", ""]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = example_fill
        row += 1

    # 빈 입력 행 200개 미리 마련
    empty_start = row
    empty_end = row + 199
    for r in range(empty_start, empty_end + 1):
        ws.cell(row=r, column=1, value=f"{r - header_row}")
        for c in range(2, 7):
            ws.cell(row=r, column=c).font = Font(name=FONT_NAME, size=10)

    # 라벨 드롭다운 데이터 검증
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(labels)}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="잘못된 라벨",
        error="드롭다운 목록에 있는 값만 선택하세요.",
    )
    ws.add_data_validation(dv)
    dv.add(f"C{header_row + 1}:C{empty_end}")

    # ambiguous 컬럼도 O/X 드롭다운
    dv2 = DataValidation(
        type="list",
        formula1='"O,X"',
        allow_blank=True,
    )
    ws.add_data_validation(dv2)
    dv2.add(f"D{header_row + 1}:D{empty_end}")

    # 컬럼 너비
    widths = {"A": 8, "B": 55, "C": 20, "D": 25, "E": 12, "F": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_stage1(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "stage1_라벨링"
    _build_sheet(ws, STAGE1_LABELS, STAGE1_EXAMPLES)
    wb.save(path)
    print(f"생성됨: {path}")


def build_stage2(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "stage2_라벨링"
    _build_sheet(ws, STAGE2_LABELS, STAGE2_EXAMPLES)
    wb.save(path)
    print(f"생성됨: {path}")


if __name__ == "__main__":
    stage1_path = os.path.join(THIS_DIR, "stage1_labeled.xlsx")
    stage2_path = os.path.join(THIS_DIR, "stage2_labeled.xlsx")

    if os.path.exists(stage1_path):
        print(f"이미 존재함, 건너뜀: {stage1_path}")
    else:
        build_stage1(stage1_path)

    if os.path.exists(stage2_path):
        print(f"이미 존재함, 건너뜀: {stage2_path}")
    else:
        build_stage2(stage2_path)
